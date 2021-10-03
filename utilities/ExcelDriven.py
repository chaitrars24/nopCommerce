import openpyxl


def getRowCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def readData(fileName, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum, column=colNum).value


def writeData(fileName, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(fileName)

